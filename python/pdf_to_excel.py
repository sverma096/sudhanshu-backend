import sys
import pdfplumber
from openpyxl import Workbook
from utils import temp_output, print_result

file_path = sys.argv[1]
wb = Workbook()
ws = wb.active
ws.title = "Extracted"

row_idx = 1
with pdfplumber.open(file_path) as pdf:
    for page in pdf.pages:
        tables = page.extract_tables() or []
        if tables:
            for table in tables:
                for row in table:
                    clean = [("" if cell is None else str(cell)) for cell in row]
                    for col_idx, val in enumerate(clean, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                    row_idx += 1
                row_idx += 1
        else:
            text = page.extract_text() or ""
            for line in text.splitlines():
                ws.cell(row=row_idx, column=1, value=line)
                row_idx += 1

out = temp_output(".xlsx")
wb.save(out)
print_result(out, "output.xlsx")
